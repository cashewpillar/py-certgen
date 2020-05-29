#! python3
# reads name from xlsx, csv, json files

import os
import openpyxl
from PIL import Image, ImageDraw, ImageFont

SOURCE = 'C:/python_files/certificate_generator/src'
DESTINATION = 'C:/python_files/certificate_generator/generated_certificates'
CERTIFICATE_TEMPLATE = 'certificate_template_1.png'
names_xlsx = 'certificate_names.xlsx'

os.chdir(SOURCE)
FONT = ImageFont.truetype('Merriweather-Bold.ttf', 42)

def generate_certificate(name, certificate_template, destination):
    # ImageFont obj used, defined beforehand
    img = Image.open(CERTIFICATE_TEMPLATE)
    d = ImageDraw.Draw(img)
    d.text((520, 580), name, fill=(0,0,0), font=FONT)
    img.save(DESTINATION + 'Certificate; ' + name + '.png')

def generate_certificates(names_xlsx):
    wb = openpyxl.load_workbook(names_xlsx)
    sheet = wb[wb.active.title]
    for r in range(1, sheet.max_row + 1):
        name = sheet[f'A{r}'].value
        generate_certificate(name, CERTIFICATE_TEMPLATE, DESTINATION)

# generate_certificates(names_xlsx)


# May 29 Trial
import subprocess, random
xl = openpyxl.open('certificate_names.xlsx')
sheet = xl.active

def openImg(filename):
    return subprocess.Popen([r'C:\Windows\System32\mspaint.exe', filename])

def generateCert(name):
    certImg = Image.open('certificate_template_1.png')
    drawCert = ImageDraw.Draw(certImg)
    #sizes
    nameWidth, nameHeight = drawCert.textsize(name, FONT)
    templateWidth, templateHeight = certImg.size
    vPlaceholder = 640
    #solving for x,y coordinates
    x = (templateWidth - nameWidth) / 2
    y = vPlaceholder - nameHeight
    #drawing text
    drawCert.text((x,y), name, fill='black', font=FONT)
    filename = 'testCert' + str(random.randint(1,9999)) + '.png'
    print(f'Generating {filename}... ')
    certImg.save(os.path.join(r'C:\python_files\certificate_generator\0529_test', filename))
    return filename

for cell in sheet.iter_rows():
    filename = generateCert(cell[0].value)
    openImg(os.path.join(r'C:\python_files\certificate_generator\0529_test', filename))
